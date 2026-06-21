#!/usr/bin/env python
"""Update CI/PL header/footer callouts from DFCR + warehouse files.

This script intentionally edits only shipment header fields and container/SKU
footer callouts. It does not rebuild middle item-detail rows.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font
from pypdf import PdfReader


MONTHS = {
    "JAN": 1,
    "FEB": 2,
    "MAR": 3,
    "APR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUNE": 6,
    "JUL": 7,
    "JULY": 7,
    "AUG": 8,
    "SEP": 9,
    "SEPT": 9,
    "OCT": 10,
    "NOV": 11,
    "DEC": 12,
}


def norm(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def compact(value: Any) -> str:
    return re.sub(r"\s+", "", norm(value))


def excel_date(dt: datetime) -> str:
    return dt.strftime("%b %d, %Y").upper()


def parse_dfcr_date(text: str) -> datetime | None:
    m = re.search(r"(\d{1,2})\s+([A-Z]+)\s+(\d{4})", text, re.I)
    if not m:
        return None
    month = MONTHS.get(m.group(2).upper())
    if not month:
        return None
    return datetime(int(m.group(3)), month, int(m.group(1)))


def read_pdf(pdf_path: Path) -> dict[str, Any]:
    text = "\n".join((page.extract_text() or "") for page in PdfReader(str(pdf_path)).pages)
    flat = re.sub(r"\s+", " ", text)

    container_match = re.search(r"CONTAINER#:(.*?)SEAL#:(.*?)(?:DRAFT|$)", flat, re.I)
    containers = re.findall(r"\b[A-Z]{4}\d{7}\b", container_match.group(1)) if container_match else []
    seals = re.findall(r"\b[A-Z]\d{7,12}\b|\b\d{8,12}\b|\b[A-Z]{4}\d{5}\b", container_match.group(2)) if container_match else []
    # Keep only likely seal tokens and preserve order.
    seals = [x for x in seals if not re.match(r"^[A-Z]{4}\d{7}$", x)]

    gross = ""
    measurement = ""
    gross_match = re.search(r"KGS\.\s*([0-9]+\.[0-9]+)\s+CBM\.\s*([0-9]+\.[0-9]+)", flat)
    if gross_match:
        gross = f"{gross_match.group(1)}KGS"
        measurement = gross_match.group(2)

    vessel = ""
    vessel_match = re.search(r"EXPORT CARRIER \(VESSEL, VOYAGE\)\s+([A-Z0-9 ]+?)\s+EXPORT REFERENCE", flat)
    if vessel_match:
        vessel = norm(vessel_match.group(1))

    etd = ""
    etd_match = re.search(r"ETD:([A-Z]+)\s+(\d{1,2}),\s*(\d{4})", flat, re.I)
    if etd_match:
        month = MONTHS.get(etd_match.group(1).upper())
        if month:
            etd = excel_date(datetime(int(etd_match.group(3)), month, int(etd_match.group(2))))

    ship_dt = None
    cargo_match = re.search(r"COUNTRY OF ORIGIN.*?(\d{1,2}\s+[A-Z][A-Z]+\s+\d{4})\s+CHINA", flat, re.I)
    if cargo_match:
        ship_dt = parse_dfcr_date(cargo_match.group(1))
    if ship_dt is None:
        first_date = re.search(r"\b\d{1,2}\s+[A-Z][A-Z]+\s+\d{4}\b", flat, re.I)
        ship_dt = parse_dfcr_date(first_date.group(0)) if first_date else None
    if ship_dt is None:
        raise RuntimeError("Could not determine SHIP DATE / cargo received date from DFCR PDF")

    destination = ""
    dest_match = re.search(r"PORT OF DISCHARGE\s+([A-Z ]+,\s*[A-Z]{2})\s+PLACE OF DELIVERY", flat)
    if dest_match:
        destination = norm(dest_match.group(1))

    return {
        "containers": containers,
        "seals": seals,
        "gross_weight": gross,
        "measurement": measurement,
        "vessel": vessel,
        "etd": etd,
        "ship_date": excel_date(ship_dt),
        "invoice_date": excel_date(ship_dt + timedelta(days=1)),
        "invoice_date_yyyymmdd": (ship_dt + timedelta(days=1)).strftime("%Y%m%d"),
        "destination": destination,
    }


def load_mapping(mapping_arg: str) -> list[dict[str, str]]:
    path = Path(mapping_arg)
    if path.exists():
        text = path.read_text(encoding="utf-8-sig")
    else:
        text = mapping_arg
    data = json.loads(text)
    if isinstance(data, dict):
        rows = []
        for pt, value in data.items():
            if isinstance(value, dict):
                rows.append({"pt": pt, "container": value["container"], "seal": value["seal"]})
            else:
                rows.append({"pt": pt, "container": value[0], "seal": value[1]})
        return rows
    return data


def run_powershell(script: str) -> None:
    with tempfile.NamedTemporaryFile("w", suffix=".ps1", encoding="utf-8-sig", delete=False) as f:
        f.write(script)
        script_path = f.name
    try:
        subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-File", script_path],
            check=True,
        )
    finally:
        try:
            os.unlink(script_path)
        except OSError:
            pass


def convert_xls_with_excel(files: list[Path], work_dir: Path) -> dict[str, Path]:
    work_dir.mkdir(parents=True, exist_ok=True)
    payload = [{"src": str(path), "dst": str(work_dir / (path.stem + ".xlsx"))} for path in files]
    json_path = work_dir / "convert_manifest.json"
    json_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    ps = rf"""
$items = Get-Content -LiteralPath '{json_path}' -Encoding UTF8 | ConvertFrom-Json
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
try {{
  foreach ($item in $items) {{
    $wb = $excel.Workbooks.Open($item.src, 0, $true)
    try {{
      $wb.SaveAs($item.dst, 51)
    }} finally {{
      $wb.Close($false)
    }}
  }}
}} finally {{
  $excel.Quit()
  [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null
}}
"""
    run_powershell(ps)
    return {Path(item["src"]): Path(item["dst"]) for item in payload}


def prepare_warehouse_files(files: list[Path], work_dir: Path) -> dict[str, Path]:
    xls_files = [path for path in files if path.suffix.lower() == ".xls"]
    converted = convert_xls_with_excel(xls_files, work_dir) if xls_files else {}
    result = {}
    for path in files:
        usable = converted.get(path, path)
        pt_match = re.search(r"PT\d{5}", path.name, re.I)
        if not pt_match:
            raise RuntimeError(f"Could not infer PT number from warehouse filename: {path}")
        result[pt_match.group(0).upper()] = usable
    return result


def read_pt_file(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    skus: list[str] = []
    pos: list[str] = []
    qty_by_sku: dict[str, int] = {}
    for row in range(9, ws.max_row + 1):
        if ws.cell(row, 1).value == "总计":
            break
        if not ws.cell(row, 1).value:
            continue
        po = compact(ws.cell(row, 17).value)
        if po and po not in pos:
            pos.append(po)
        sku = compact(ws.cell(row, 18).value) or compact(ws.cell(row, 19).value)
        if sku and sku not in skus:
            skus.append(sku)
        qty = ws.cell(row, 22).value
        if sku and isinstance(qty, (int, float)):
            qty_by_sku[sku] = qty_by_sku.get(sku, 0) + int(qty)
    return {"skus": skus, "pos": pos, "qty_by_sku": qty_by_sku}


def ordered_container_data(pdf: dict[str, Any], mapping: list[dict[str, str]], warehouse_by_pt: dict[str, Path]) -> list[dict[str, Any]]:
    pt_by_pair = {(row["container"], row["seal"]): row["pt"].upper() for row in mapping}
    pt_by_seal = {row["seal"]: row["pt"].upper() for row in mapping}
    ordered = []
    for container, seal in zip(pdf["containers"], pdf["seals"]):
        pt = pt_by_pair.get((container, seal)) or pt_by_seal.get(seal)
        if not pt:
            raise RuntimeError(f"No PT mapping found for {container}/{seal}")
        if pt not in warehouse_by_pt:
            raise RuntimeError(f"No warehouse file was supplied for {pt}")
        pt_data = read_pt_file(warehouse_by_pt[pt])
        ordered.append({"container": container, "seal": seal, "pt": pt, **pt_data})
    return ordered


def build_callout(ordered: list[dict[str, Any]]) -> str:
    lines = ["CONTAINER/SEAL NO.:"]
    locations: dict[str, list[dict[str, Any]]] = {}
    for idx, item in enumerate(ordered):
        end = "." if idx == len(ordered) - 1 else ";"
        lines.append(f"{item['container']}/{item['seal']} including sku#{', '.join(item['skus'])}{end}")
        for sku in item["skus"]:
            locations.setdefault(sku, []).append(item)
    for sku, items in locations.items():
        if len(items) <= 1:
            continue
        if len(items) == 2:
            first, second = items
            lines.append(
                f"SKU# {sku}: {first['qty_by_sku'].get(sku, 0)} pcs were loaded into container {first['container']}; "
                f"the remaining {second['qty_by_sku'].get(sku, 0)} pcs were loaded into container {second['container']}."
            )
        else:
            parts = [f"{item['qty_by_sku'].get(sku, 0)} pcs were loaded into container {item['container']}" for item in items]
            lines.append(f"SKU# {sku}: " + "; ".join(parts) + ".")
    return "\n".join(lines)


def unique_po_text(ordered: list[dict[str, Any]]) -> str:
    pos: list[str] = []
    for item in ordered:
        for po in item["pos"]:
            if po not in pos:
                pos.append(po)
    return ", ".join(pos)


def find_invoice_no(wb) -> str:
    value = str(wb["CI "]["H2"].value or "")
    match = re.search(r"INVOICE\s*No\.?:\s*([A-Z0-9-]+)", value, re.I)
    return match.group(1) if match else "ZSEFP-2600XXX"


def mask_invoice_no(invoice_no: str) -> str:
    return re.sub(r"[A-Z0-9]{3}$", "XXX", invoice_no.upper())


def red_font_like(cell) -> Font:
    return Font(
        name=cell.font.name,
        sz=cell.font.sz,
        bold=cell.font.bold,
        italic=cell.font.italic,
        vertAlign=cell.font.vertAlign,
        underline=cell.font.underline,
        strike=cell.font.strike,
        color="FF0000",
    )


def update_workbook(template: Path, output: Path, pdf: dict[str, Any], ordered: list[dict[str, Any]], callout: str) -> str:
    shutil.copy2(template, output)
    wb = load_workbook(output)
    invoice_no = find_invoice_no(wb)
    display_invoice_no = mask_invoice_no(invoice_no)
    ci = wb["CI "]
    pl = wb["PL"]

    ci["J7"] = f"SHIP DATE:                                                 \n{pdf['ship_date']}"
    ci["J10"] = f"ETD\n{pdf['etd']}"
    ci["J12"] = f"TOTAL GROSS WEIGHT:\n{pdf['gross_weight']}"
    if ci["H12"].value:
        ci["H12"].font = red_font_like(ci["H12"])

    # Vendor and factory templates can place this header block on different rows.
    if isinstance(ci["E14"].value, str) and ci["E14"].value.startswith("CONTAINER/SEAL NO.:"):
        container_cell, destination_cell, vessel_cell, sailing_cell = "E14", "F14", "E16", "F16"
    else:
        container_cell, destination_cell, vessel_cell, sailing_cell = "E16", "F16", "E18", "F18"

    ci[container_cell] = "CONTAINER/SEAL NO.:\n" + "\n".join(
        f"{item['container']}/{item['seal']}{';' if idx < len(ordered)-1 else '.'}"
        for idx, item in enumerate(ordered)
    )
    ci[destination_cell] = f"FINAL DESTINATION\n{pdf['destination']}"
    ci[vessel_cell] = f"VESSEL: \n{pdf['vessel']}"
    ci[sailing_cell] = f"SAILING ON OR ABOUT :\n{pdf['etd']}"

    for row in ci.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("CONTAINER/SEAL NO.:") and cell.coordinate != container_cell:
                cell.value = callout
                cell.alignment = cell.alignment.copy(wrap_text=True, vertical="top")
                ci.row_dimensions[cell.row].height = max(ci.row_dimensions[cell.row].height or 0, 88)
                if cell.row + 1 <= ci.max_row:
                    ci.row_dimensions[cell.row + 1].height = max(ci.row_dimensions[cell.row + 1].height or 0, 50)

    pl["A4"] = (
        f"CUSTOMER PO NO.: {unique_po_text(ordered)}.\n"
        f"INVOICE NO.: {display_invoice_no}\n"
        f"INVOICE DATE: {pdf['invoice_date']}                                                                                                               \n"
        f"SHIPPING METHOD: BY SEA                                                                                                                                                                                               ETD: {pdf['etd']}                                                                                                                                                                                        \n"
        f"PORT OF DISCHARGE: {pdf['destination']}\n"
        "TERMS OF SALES PAYMENT AND DISCOUNT REFERENCE: FOB, FREIGHT COLLECT"
    )
    for row in pl.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("CONTAINER/SEAL NO.:"):
                cell.value = callout
                cell.alignment = cell.alignment.copy(wrap_text=True, vertical="top")
                pl.row_dimensions[cell.row].height = max(pl.row_dimensions[cell.row].height or 0, 88)

    wb.save(output)
    return invoice_no


def rich_text_finalize(outputs: list[Path], pdf: dict[str, Any], mark_invoice_red: bool) -> None:
    payload = [{"file": str(path), "invoice_date": pdf["invoice_date"], "mark_invoice_red": mark_invoice_red} for path in outputs]
    with tempfile.NamedTemporaryFile("w", suffix=".json", encoding="utf-8", delete=False) as f:
        json.dump(payload, f, ensure_ascii=False)
        payload_path = f.name
    ps = rf"""
$items = Get-Content -LiteralPath '{payload_path}' -Encoding UTF8 | ConvertFrom-Json
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
try {{
  foreach ($item in $items) {{
    $wb = $excel.Workbooks.Open($item.file)
    try {{
      $ws = $wb.Worksheets.Item('CI ')
      $cell = $ws.Range('H2')
      $text = [string]$cell.Value2
      $lineBreak = $text.IndexOf("`n")
      $dateMatch = [regex]::Match($text, '(?i)(INVOICE DATE:\s*)((?:JANUARY|FEBRUARY|MARCH|APRIL|MAY|JUNE|JULY|AUGUST|SEPTEMBER|OCTOBER|NOVEMBER|DECEMBER|JAN|FEB|MAR|APR|JUN|JUL|AUG|SEP|OCT|NOV|DEC)\s+\d{{1,2}},\s+\d{{4}})')
      if ($dateMatch.Success) {{
        $start = $dateMatch.Groups[2].Index + 1
        $len = $dateMatch.Groups[2].Length
        $cell.Characters($start, $len).Text = $item.invoice_date
      }}
      $text = [string]$cell.Value2
      $lineBreak = $text.IndexOf("`n")
      if ($lineBreak -ge 0) {{
        $cell.Characters(1, $lineBreak).Font.Size = 14
        $cell.Characters($lineBreak + 2, $text.Length - $lineBreak - 1).Font.Size = 16
      }}
      if ($item.mark_invoice_red) {{
        $invMatch = [regex]::Match([string]$cell.Value2, '(?im)(INVOICE\s*No\.?:\s*)(.+)$')
        if ($invMatch.Success) {{
          $start = $invMatch.Groups[2].Index + 1
          $masked = [regex]::Replace($invMatch.Groups[2].Value.ToUpper(), '[A-Z0-9]{{3}}$', 'XXX')
          $cell.Characters($start, $invMatch.Groups[2].Length).Text = $masked
          $len = $masked.Length
          $cell.Characters($start, $len).Font.Color = 255
        }}
        $pl = $wb.Worksheets.Item('PL')
        $plCell = $pl.Range('A4')
        $plInvMatch = [regex]::Match([string]$plCell.Value2, '(?im)(INVOICE\s*NO\.?:\s*)(.+)$')
        if ($plInvMatch.Success) {{
          $start = $plInvMatch.Groups[2].Index + 1
          $masked = [regex]::Replace($plInvMatch.Groups[2].Value.ToUpper(), '[A-Z0-9]{{3}}$', 'XXX')
          $plCell.Characters($start, $plInvMatch.Groups[2].Length).Text = $masked
          $len = $masked.Length
          $plCell.Characters($start, $len).Font.Color = 255
        }}
      }}
      $wb.Save()
    }} finally {{
      $wb.Close($true)
    }}
  }}
}} finally {{
  $excel.Quit()
  [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null
}}
"""
    try:
        run_powershell(ps)
    finally:
        try:
            os.unlink(payload_path)
        except OSError:
            pass


def main() -> int:
    parser = argparse.ArgumentParser(description="Update CI/PL header/footer fields from DFCR + warehouse files.")
    parser.add_argument("--dfcr", required=True, type=Path)
    parser.add_argument("--mapping", required=True, help="JSON string or path. Rows: {container, seal, pt}.")
    parser.add_argument("--warehouse", required=True, nargs="+", type=Path, help="Warehouse XLS/XLSX files.")
    parser.add_argument("--vendor-template", type=Path)
    parser.add_argument("--factory-template", type=Path)
    parser.add_argument("--output-dir", type=Path, default=Path.cwd())
    parser.add_argument("--work-dir", type=Path, default=Path(tempfile.gettempdir()) / "ci_pl_update_work")
    parser.add_argument("--no-mark-invoice-red", action="store_true")
    parser.add_argument("--summary-json", type=Path)
    args = parser.parse_args()

    templates = [p for p in [args.vendor_template, args.factory_template] if p]
    if not templates:
        raise SystemExit("Provide --vendor-template and/or --factory-template")

    args.output_dir.mkdir(parents=True, exist_ok=True)
    pdf = read_pdf(args.dfcr)
    mapping = load_mapping(args.mapping)
    warehouse_by_pt = prepare_warehouse_files(args.warehouse, args.work_dir)
    ordered = ordered_container_data(pdf, mapping, warehouse_by_pt)
    callout = build_callout(ordered)

    outputs = []
    invoices = {}
    for template in templates:
        out_name = template.name
        template_wb = load_workbook(template, read_only=True, data_only=False)
        try:
            template_invoice_no = find_invoice_no(template_wb)
        finally:
            template_wb.close()
        if not args.no_mark_invoice_red:
            out_name = re.sub(re.escape(template_invoice_no), mask_invoice_no(template_invoice_no), out_name, flags=re.I)
        fcr_match = re.search(r"NGB\d+", args.dfcr.name, re.I)
        if fcr_match:
            out_name = re.sub(r"NGB\d+", fcr_match.group(0).upper(), out_name, flags=re.I)
        out_name = re.sub(r"20\d{6}", pdf["invoice_date_yyyymmdd"], out_name, count=1)
        output = args.output_dir / out_name
        invoice_no = update_workbook(template, output, pdf, ordered, callout)
        invoices[str(output)] = invoice_no
        outputs.append(output)

    rich_text_finalize(outputs, pdf, mark_invoice_red=not args.no_mark_invoice_red)

    summary = {
        "outputs": [str(p) for p in outputs],
        "pdf": pdf,
        "order": [{"container": x["container"], "seal": x["seal"], "pt": x["pt"]} for x in ordered],
        "callout": callout,
        "invoice_numbers": invoices,
    }
    text = json.dumps(summary, ensure_ascii=False, indent=2)
    if args.summary_json:
        args.summary_json.write_text(text, encoding="utf-8")
    print(text)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
